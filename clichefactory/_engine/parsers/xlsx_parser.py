from io import BytesIO
import re
from typing import List, Optional, Tuple
import logging

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries


from clichefactory._engine.models.document_model import Block, Heading, Paragraph, Table, TableCell
from clichefactory._engine.models.normalized_doc import NormalizedDoc
from clichefactory._engine.adapters.xlsx_adapter import XlsxNormalizedDoc
from clichefactory._engine.parsers.media_parser import MediaParser


logger = logging.getLogger(__name__)


def _is_empty(v: object) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _normalize_cell_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v)


def _sheet_used_bounds(ws: Worksheet) -> Optional[Tuple[int, int, int, int]]:
    """
    Return (min_row, min_col, max_row, max_col) for non-empty cells.
    openpyxl's ws.calculate_dimension() is loose; this scans quickly.
    """
    min_r, min_c = None, None
    max_r, max_c = None, None

    # Restrict scan to ws.max_row/max_column (openpyxl computed)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if not _is_empty(ws.cell(row=r, column=c).value):
                min_r = r if min_r is None else min(min_r, r)
                max_r = r if max_r is None else max(max_r, r)
                min_c = c if min_c is None else min(min_c, c)
                max_c = c if max_c is None else max(max_c, c)
        # small optimization: if we're past max_r by a lot, you could break,
        # but for most sheets ws.max_row isn't huge.
    if min_r is None:
        return None
    return (min_r, min_c or 1, max_r or min_r, max_c or min_c or 1)


def _row_stats(ws: Worksheet, row: int, min_col: int, max_col: int) -> dict:
    values = [ws.cell(row=row, column=c).value for c in range(min_col, max_col + 1)]
    non_empty = [v for v in values if not _is_empty(v)]
    num_values = len(values)
    num_non_empty = len(non_empty)

    num_str = sum(isinstance(v, str) for v in non_empty)
    num_num = sum(isinstance(v, (int, float)) for v in non_empty)
    num_date = sum(hasattr(v, "year") and hasattr(v, "month") for v in non_empty)  # naive date-ish

    # style signals
    bold = 0
    filled = 0
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.value is None:
            continue
        if cell.font is not None and bool(getattr(cell.font, "bold", False)):
            bold += 1
        # patternType or fgColor set often indicates a header fill
        fill = cell.fill
        if fill is not None and getattr(fill, "patternType", None):
            filled += 1

    return {
        "num_values": num_values,
        "num_non_empty": num_non_empty,
        "density": (num_non_empty / num_values) if num_values else 0.0,
        "num_str": num_str,
        "num_num": num_num,
        "num_date": num_date,
        "num_bold": bold,
        "num_filled": filled,
    }



def detect_header_row(ws: Worksheet, bounds: Tuple[int, int, int, int]) -> Optional[int]:
    """
    Returns 1-based row index of header, or None if no good header.
    """
    min_row, min_col, max_row, max_col = bounds

    # 1) Freeze panes heuristic (very strong)
    if ws.freeze_panes is not None:
        # freeze_panes is a cell like "A2" meaning rows above are frozen => header likely row 1
        try:
            ref = ws.freeze_panes.coordinate  # type: ignore[attr-defined]
        except Exception:
            ref = str(ws.freeze_panes)

        m = re.match(r"([A-Z]+)(\d+)", ref)
        if m:
            freeze_row = int(m.group(2))
            candidate = freeze_row - 1
            if candidate >= min_row:
                return candidate

    # 2) Excel Table object heuristic (strong)
    # openpyxl stores tables in ws.tables (dict-like)
    try:
        if ws.tables:
            # take first table; for multiple, you can choose the top-most
            t = next(iter(ws.tables.values()))
            # table.ref e.g. "A1:D20" typically includes header row at top
            min_c, min_r, max_c, max_r = range_boundaries(t.ref)
            return min_r
    except Exception:
        logger.warning(f"Exception thrown when trying to determine header. row")

    # 3) Score-based heuristic over the first ~50 rows of used range
    scan_to = min(max_row, min_row + 50)

    best_row = None
    best_score = float("-inf")

    for r in range(min_row, scan_to + 1):
        s = _row_stats(ws, r, min_col, max_col)
        if s["num_non_empty"] == 0:
            continue

        # next row stats to detect header->data transition
        s_next = _row_stats(ws, r + 1, min_col, max_col) if r + 1 <= max_row else None

        score = 0.0

        # dense row is more likely header
        score += 2.0 * s["density"]

        # header rows tend to be mostly strings
        if s["num_non_empty"] > 0:
            score += 1.5 * (s["num_str"] / s["num_non_empty"])
            score -= 0.5 * (s["num_num"] / s["num_non_empty"])
            score -= 0.5 * (s["num_date"] / s["num_non_empty"])

        # styling: bold/fill are good hints
        score += 0.8 * (s["num_bold"] / max(1, s["num_non_empty"]))
        score += 0.5 * (s["num_filled"] / max(1, s["num_non_empty"]))

        # transition: next row more numeric/date than current suggests current is header
        if s_next and s_next["num_non_empty"] > 0:
            next_num_ratio = (s_next["num_num"] + s_next["num_date"]) / s_next["num_non_empty"]
            cur_num_ratio = (s["num_num"] + s["num_date"]) / s["num_non_empty"]
            if next_num_ratio > cur_num_ratio:
                score += 0.8

        # penalize if row looks like a title row (single cell)
        if s["num_non_empty"] == 1 and s["density"] < 0.3:
            score -= 1.5

        if score > best_score:
            best_score = score
            best_row = r

    # Require some confidence
    if best_row is None:
        return None
    if best_score < 1.2:
        return None
    return best_row


def sheet_to_markdown_table(ws: Worksheet, bounds: Tuple[int, int, int, int], header_row: Optional[int]) -> Tuple[str, Table]:
    min_row, min_col, max_row, max_col = bounds

    # decide header + data start
    if header_row is None:
        header_row = min_row
        data_start = min_row
        has_header = False
    else:
        data_start = header_row + 1
        has_header = True

    # Build rows of strings
    def row_values(r: int) -> List[str]:
        return [_normalize_cell_text(ws.cell(row=r, column=c).value) for c in range(min_col, max_col + 1)]

    header = row_values(header_row)
    body_rows: List[List[str]] = []

    for r in range(data_start, max_row + 1):
        vals = row_values(r)
        if all(v == "" for v in vals):
            # you can decide to stop at first fully empty row; often OK for tables
            continue
        body_rows.append(vals)

    # Trim trailing completely empty columns across header+body
    all_rows = [header] + body_rows
    if all_rows:
        keep_cols = []
        for j in range(len(all_rows[0])):
            if any((row[j].strip() if j < len(row) else "") for row in all_rows):
                keep_cols.append(j)
        header = [header[j] for j in keep_cols]
        body_rows = [[row[j] for j in keep_cols] for row in body_rows]
        max_col_effective = min_col + len(keep_cols) - 1
    else:
        max_col_effective = max_col

    # Markdown render
    def md_row(cols: List[str]) -> str:
        safe = [c.replace("\n", " ").strip() for c in cols]
        return "| " + " | ".join(safe) + " |"

    if not header:
        md = ""
    else:
        if has_header:
            md_lines = [md_row(header), "| " + " | ".join(["---"] * len(header)) + " |"]
        else:
            # no header: create a synthetic one
            synth = [f"col_{i+1}" for i in range(len(header))]
            md_lines = [md_row(synth), "| " + " | ".join(["---"] * len(synth)) + " |"]
            # treat original "header" row as first data row
            body_rows = [header] + body_rows

        for r in body_rows:
            md_lines.append(md_row(r))
        md = "\n".join(md_lines)

    # Build document_model.Table cells (row/col are 0-based in your model usage)
    cells: List[TableCell] = []
    # header row as row=0, body continues
    row0 = 0
    if has_header:
        for j, txt in enumerate(header):
            cells.append(TableCell(text=txt, row=row0, col=j, bbox=None))
        start_row_idx = 1
        for i, r in enumerate(body_rows):
            for j, txt in enumerate(r):
                cells.append(TableCell(text=txt, row=start_row_idx + i, col=j, bbox=None))
    else:
        # we used synthetic header; only store data rows as-is (or store synth header too if you want)
        for j, txt in enumerate([f"col_{i+1}" for i in range(len(header))]):
            cells.append(TableCell(text=txt, row=0, col=j, bbox=None))
        for i, r in enumerate(body_rows):
            for j, txt in enumerate(r):
                cells.append(TableCell(text=txt, row=1 + i, col=j, bbox=None))

    return md, Table(cells=tuple(cells), bbox=None)

  
class XlsxParser(MediaParser):
    """
    XLSX -> openpyxl Workbook -> NormalizedDoc
    Creates:
      H1 = workbook title
      H2 = sheet name
      Table block per sheet (markdown table in get_markdown)
    """

    def __init__(self, cacher=None) -> None:
        super().__init__(cacher=cacher)

    def document_parse(self, content: bytes, filename: str) -> NormalizedDoc:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

        workbook_title = filename.rsplit(".", 1)[0] if filename else "Workbook"

        blocks: List[Block] = [Heading(level=1, text=workbook_title)]
        md_lines: List[str] = [f"# {workbook_title}"]

        for ws in wb.worksheets:
            blocks.append(Heading(level=2, text=ws.title))
            md_lines.append("")  # blank line
            md_lines.append(f"## {ws.title}")

            bounds = _sheet_used_bounds(ws)
            if not bounds:
                blocks.append(Paragraph(text="_Empty sheet_"))
                md_lines.append("")
                md_lines.append("_Empty sheet_")
                continue

            header_row = detect_header_row(ws, bounds)
            sheet_md, table_block = sheet_to_markdown_table(ws, bounds, header_row)

            # If you want the exact md per sheet, you can store it as Paragraph
            # and/or rely on Table->md reconstruction.
            # Here we add the Table block so you can access cells structurally.
            blocks.append(table_block)

            # markdown representation
            if sheet_md:
                md_lines.append("")
                md_lines.append(sheet_md)

        markdown = "\n".join(md_lines).strip() + "\n"

        return XlsxNormalizedDoc(filename=filename, sheet_blocks=blocks, markdown=markdown)
